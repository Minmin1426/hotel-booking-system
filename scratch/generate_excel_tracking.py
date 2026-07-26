import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def build_excel_tracking():
    wb = openpyxl.Workbook()
    
    # Styles
    title_font = Font(name="Calibri", size=16, bold=True, color="1F497D")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    
    sub_header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    
    section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    section_font = Font(name="Calibri", size=12, bold=True, color="1F497D")
    
    completed_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    completed_font = Font(name="Calibri", size=10, bold=True, color="375623")
    
    regular_font = Font(name="Calibri", size=10)
    bold_font = Font(name="Calibri", size=10, bold=True)
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    # -------------------------------------------------------------
    # Sheet 1: Dashboard & Overview
    # -------------------------------------------------------------
    ws1 = wb.active
    ws1.title = "Dashboard & Overview"
    ws1.views.sheetView[0].showGridLines = True
    
    ws1.cell(row=1, column=1, value="DỰ ÁN HỆ THỐNG ĐẶT PHÒNG KHÁCH SẠN MỞ RỘNG (HOTEL BOOKING SYSTEM)").font = title_font
    ws1.cell(row=2, column=1, value="BÁO CÁO TIẾN ĐỘ & DỮ LIỆU DASHBOARD PROJECT TRACKING MASTER").font = Font(name="Calibri", size=12, italic=True, color="595959")
    
    dashboard_headers = ["Hạng Mục Quản Lý", "Chỉ Số / Kết Quả", "Tỷ Lệ Hoàn Thành", "Ghi Chú & Chi Tiết Triển Khai"]
    for col_idx, h in enumerate(dashboard_headers, 1):
        cell = ws1.cell(row=4, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    dashboard_data = [
        ("Tổng Số Màn Hình System (UI)", "50 / 50 Screens", "100%", "5 Phân hệ, 5 Role (Mỗi role 10 màn hình)"),
        ("Nghiệp Vụ Use Cases (SDD Spec)", "35 / 35 Use Cases", "100%", "Bao phủ 10 Feature Packages Spring Boot"),
        ("Backend Integration & Unit Tests", "207 / 207 Passed", "100%", "0 Failures, 0 Errors (Maven test JDK 18)"),
        ("Frontend Production Build", "Build Success (1.79s)", "100%", "Vite 8 + React 18, Bundle size 496 kB"),
        ("Database Flyway Migrations", "V1 ➔ V30 Migrated", "100%", "PostgreSQL Neon Cloud DB & SQL Server"),
        ("Containerization (Docker)", "Full-Stack Dockerized", "100%", "Dockerfile, frontend/Dockerfile, docker-compose.yml"),
        ("Git Remote Repository Sync", "Commit 2268f24", "100%", "Synced with origin/main (GitHub)")
    ]
    
    for row_offset, row_data in enumerate(dashboard_data, 5):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws1.cell(row=row_offset, column=col_idx, value=val)
            cell.font = bold_font if col_idx <= 2 else regular_font
            cell.border = thin_border
            if col_idx == 3:
                cell.fill = completed_fill
                cell.font = completed_font
                cell.alignment = Alignment(horizontal="center")

    # -------------------------------------------------------------
    # Sheet 2: 50 Screens Matrix
    # -------------------------------------------------------------
    ws2 = wb.create_sheet(title="50 Screens Matrix")
    ws2.views.sheetView[0].showGridLines = True
    
    ws2.cell(row=1, column=1, value="MA TRẬN TIẾN ĐỘ 50 MÀN HÌNH (SCR-101 đến SCR-510)").font = title_font
    
    screen_headers = ["Mã Screen", "Phân Hệ / Module", "Tên Màn Hình & Nghiệp Vụ", "Vai Trò (Role)", "URL Route Frontend", "API Component Backend", "Trạng Thái"]
    for col_idx, h in enumerate(screen_headers, 1):
        cell = ws2.cell(row=3, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    screens_list = [
        # Phân hệ 1
        ("SCR-101", "Phân Hệ 1: Cổng Khách Hàng", "Đăng ký, Đăng nhập & Google OAuth", "Guest / Customer", "/login, /register", "AuthController.java", "COMPLETED"),
        ("SCR-102", "Phân Hệ 1: Cổng Khách Hàng", "Hồ sơ Cá nhân & Doanh nghiệp CTP", "Customer", "/profile?tab=ctp", "UserController.java", "COMPLETED"),
        ("SCR-103", "Phân Hệ 1: Cổng Khách Hàng", "Quản lý Danh sách Đoàn & Excel Import", "Customer / Organizer", "/hotels/:id (Manifest)", "CustomerPortalController.java", "COMPLETED"),
        ("SCR-104", "Phân Hệ 1: Cổng Khách Hàng", "Khách hàng Thân thiết & Hạng Loyalty", "Customer", "/profile?tab=loyalty", "UserController.java", "COMPLETED"),
        ("SCR-105", "Phân Hệ 1: Cổng Khách Hàng", "Ví Điện Tử Cá Nhân & Số dư Ví Đoàn", "Customer", "/profile?tab=wallet", "PaymentController.java", "COMPLETED"),
        ("SCR-106", "Phân Hệ 1: Cổng Khách Hàng", "Nạp tiền Ví & Hạn mức Chi tiêu Ngày", "Customer", "/profile?tab=wallet", "PaymentController.java", "COMPLETED"),
        ("SCR-107", "Phân Hệ 1: Cổng Khách Hàng", "Kho Vé ăn & Mã QR Code Suất ăn", "Customer", "/profile?tab=mealtickets", "CustomerPortalController.java", "COMPLETED"),
        ("SCR-108", "Phân Hệ 1: Cổng Khách Hàng", "Yêu cầu Hủy đơn & Dự toán Hoàn tiền", "Customer", "/profile?tab=bookings", "BookingController.java", "COMPLETED"),
        ("SCR-109", "Phân Hệ 1: Cổng Khách Hàng", "Kho Voucher & Đổi điểm Thưởng", "Customer", "/profile?tab=vouchers", "VoucherController.java", "COMPLETED"),
        ("SCR-110", "Phân Hệ 1: Cổng Khách Hàng", "Admin View Quản lý Người dùng & Đoàn", "Admin", "/admin/users?tab=users", "UserController.java", "COMPLETED"),
        
        # Phân hệ 2
        ("SCR-201", "Phân Hệ 2: Vận Hành & Ma Trận", "Dashboard Tổng quan Đối tác Khách sạn", "Hotel Manager", "/admin/users?tab=hotels", "HotelController.java", "COMPLETED"),
        ("SCR-202", "Phân Hệ 2: Vận Hành & Ma Trận", "Khai báo Khách sạn & Khu vực Nhà hàng", "Hotel Manager", "/admin/users?tab=hotels (Modal)", "HotelController.java", "COMPLETED"),
        ("SCR-203", "Phân Hệ 2: Vận Hành & Ma Trận", "Quản lý Loại phòng & Quỹ phòng Đoàn", "Hotel Manager", "/admin/users?tab=rooms", "RoomController.java", "COMPLETED"),
        ("SCR-204", "Phân Hệ 2: Vận Hành & Ma Trận", "Sơ đồ Phòng Real-time (Room Matrix)", "Housekeeper / Reception", "/staff/rooms", "RoomController.java", "COMPLETED"),
        ("SCR-205", "Phân Hệ 2: Vận Hành & Ma Trận", "Phân bổ Phòng Hàng loạt cho Đoàn", "Receptionist", "/hotels/:id (Manifest)", "BookingController.java", "COMPLETED"),
        ("SCR-206", "Phân Hệ 2: Vận Hành & Ma Trận", "Quản lý Menu Nhà hàng & Gói Vé ăn", "Restaurant Staff", "/hotels/:id?tab=meal", "CustomerPortalController.java", "COMPLETED"),
        ("SCR-207", "Phân Hệ 2: Vận Hành & Ma Trận", "Quét Mã QR Vé ăn tại Nhà hàng", "Restaurant Staff", "/staff/rooms (Scanner)", "CustomerPortalController.java", "COMPLETED"),
        ("SCR-208", "Phân Hệ 2: Vận Hành & Ma Trận", "Cấu hình Giá phòng & Chiết khấu Đoàn", "Hotel Manager", "/admin/users?tab=rooms", "RoomController.java", "COMPLETED"),
        ("SCR-209", "Phân Hệ 2: Vận Hành & Ma Trận", "Duyệt Yêu cầu Hủy phòng & Hoàn tiền", "Admin / Finance", "/admin/users?tab=bookings", "BookingController.java", "COMPLETED"),
        ("SCR-210", "Phân Hệ 2: Vận Hành & Ma Trận", "Admin View Duyệt Khách sạn mới", "Admin", "/admin/users?tab=hotels", "HotelController.java", "COMPLETED"),

        # Phân hệ 3
        ("SCR-301", "Phân Hệ 3: Đặt Đoàn & Lễ Tân", "Tìm kiếm Khách sạn Lẻ & Đoàn", "Customer / Organizer", "/ (Hero Search)", "HotelController.java", "COMPLETED"),
        ("SCR-302", "Phân Hệ 3: Đặt Đoàn & Lễ Tân", "Kết quả Tìm kiếm & Bộ lọc Combo", "Customer", "/ (Catalog Grid)", "HotelController.java", "COMPLETED"),
        ("SCR-303", "Phân Hệ 3: Đặt Đoàn & Lễ Tân", "Chi tiết Khách sạn & Menu Nhà hàng", "Customer", "/hotels/:id", "HotelController.java", "COMPLETED"),
        ("SCR-304", "Phân Hệ 3: Đặt Đoàn & Lễ Tân", "Wizard Đặt phòng Đoàn (Group Booking)", "Organizer", "/hotels/:id?tab=group", "BookingController.java", "COMPLETED"),
        ("SCR-305", "Phân Hệ 3: Đặt Đoàn & Lễ Tân", "Đặt bàn & Mua Vé ăn Riêng lẻ", "Customer", "/hotels/:id?tab=meal", "CustomerPortalController.java", "COMPLETED"),
        ("SCR-306", "Phân Hệ 3: Đặt Đoàn & Lễ Tân", "Nhập Danh sách Khách Đoàn & Phân công", "Organizer", "/hotels/:id (Manifest)", "CustomerPortalController.java", "COMPLETED"),
        ("SCR-307", "Phân Hệ 3: Đặt Đoàn & Lễ Tân", "Chi tiết Đơn Đặt đoàn & Mã QR Tổng", "Customer", "/hotels/:id (Checkout)", "BookingController.java", "COMPLETED"),
        ("SCR-308", "Phân Hệ 3: Đặt Đoàn & Lễ Tân", "Lễ tân Quản lý & Tiếp nhận Đơn Đoàn", "Receptionist", "/staff/rooms", "BookingController.java", "COMPLETED"),
        ("SCR-309", "Phân Hệ 3: Đặt Đoàn & Lễ Tân", "Lễ tân Check-in Đoàn Cấp tốc", "Receptionist", "/staff/rooms (Express)", "BookingController.java", "COMPLETED"),
        ("SCR-310", "Phân Hệ 3: Đặt Đoàn & Lễ Tân", "Lễ tân Check-out Đoàn & Phụ thu", "Receptionist", "/staff/rooms", "BookingController.java", "COMPLETED"),

        # Phân hệ 4
        ("SCR-401", "Phân Hệ 4: Cọc Đoàn & VAT", "Chọn Phương thức Thanh toán Combo", "Customer", "/hotels/:id (Checkout)", "PaymentController.java", "COMPLETED"),
        ("SCR-402", "Phân Hệ 4: Cọc Đoàn & VAT", "Thanh toán Đặt cọc Đoàn (30% Deposit)", "Organizer", "/hotels/:id?tab=group", "PaymentController.java", "COMPLETED"),
        ("SCR-403", "Phân Hệ 4: Cọc Đoàn & VAT", "Kết quả Thanh toán & Biên lai Điện tử", "Customer", "/payment/success", "PaymentController.java", "COMPLETED"),
        ("SCR-404", "Phân Hệ 4: Cọc Đoàn & VAT", "Engine & Màn hình Hoàn tiền Tự động", "Customer / System", "/profile?tab=bookings", "BookingController.java", "COMPLETED"),
        ("SCR-405", "Phân Hệ 4: Cọc Đoàn & VAT", "Xử lý Hoàn tiền Vé ăn Thừa", "Customer", "/profile?tab=bookings", "BookingController.java", "COMPLETED"),
        ("SCR-406", "Phân Hệ 4: Cọc Đoàn & VAT", "Quản lý Mã Giảm giá Combo", "Admin", "/admin/users?tab=vouchers", "VoucherController.java", "COMPLETED"),
        ("SCR-407", "Phân Hệ 4: Cọc Đoàn & VAT", "Tra cứu Dòng tiền Cọc & Lịch sử Giao dịch", "Finance Admin", "/profile?tab=wallet", "PaymentController.java", "COMPLETED"),
        ("SCR-408", "Phân Hệ 4: Cọc Đoàn & VAT", "Xuất Hóa đơn Red VAT Doanh nghiệp", "Finance Admin", "/admin/users?tab=ctp", "CustomerPortalController.java", "COMPLETED"),
        ("SCR-409", "Phân Hệ 4: Cọc Đoàn & VAT", "Đối soát Doanh thu Phòng & Nhà hàng", "Finance Admin", "/admin/users?tab=reports", "ReportController.java", "COMPLETED"),
        ("SCR-410", "Phân Hệ 4: Cọc Đoàn & VAT", "Admin Payout & Duyệt lệnh Hoàn tiền", "Finance Admin", "/admin/users?tab=bookings", "PaymentController.java", "COMPLETED"),

        # Phân hệ 5
        ("SCR-501", "Phân Hệ 5: Real-time Chat & Analytics", "Chat Real-time Khách <-> Lễ tân", "Customer / Reception", "Header & /staff/rooms", "StaffChatController.java", "COMPLETED"),
        ("SCR-502", "Phân Hệ 5: Real-time Chat & Analytics", "Chat Hỗ trợ Trưởng đoàn Dedicated", "Organizer / Staff", "/staff/rooms", "StaffChatController.java", "COMPLETED"),
        ("SCR-503", "Phân Hệ 5: Real-time Chat & Analytics", "AI Chatbot Tư vấn Combo & Tiệc Đoàn", "Guest / Customer", "/ (Floating AI Widget)", "AIChatController.java", "COMPLETED"),
        ("SCR-504", "Phân Hệ 5: Real-time Chat & Analytics", "Order Dịch vụ Tận phòng (Room Service)", "Customer / Staff", "/staff/rooms", "StaffRoomController.java", "COMPLETED"),
        ("SCR-505", "Phân Hệ 5: Real-time Chat & Analytics", "Quản lý Thông báo Push Notification", "Customer", "Header Badge Notifications", "NotificationController.java", "COMPLETED"),
        ("SCR-506", "Phân Hệ 5: Real-time Chat & Analytics", "Dashboard Báo cáo Doanh thu Giám đốc", "Director", "/admin/users?tab=reports", "ReportController.java", "COMPLETED"),
        ("SCR-507", "Phân Hệ 5: Real-time Chat & Analytics", "Báo cáo Khách đoàn vs Khách lẻ", "Director", "/admin/users?tab=reports", "ReportController.java", "COMPLETED"),
        ("SCR-508", "Phân Hệ 5: Real-time Chat & Analytics", "Báo cáo Hủy phòng & Chi phí Hoàn tiền", "Director", "/admin/users?tab=reports", "ReportController.java", "COMPLETED"),
        ("SCR-509", "Phân Hệ 5: Real-time Chat & Analytics", "Báo cáo Doanh thu Nhà hàng & Vé ăn", "Director", "/admin/users?tab=reports", "ReportController.java", "COMPLETED"),
        ("SCR-510", "Phân Hệ 5: Real-time Chat & Analytics", "Xuất Báo cáo Excel / PDF Tổng hợp", "Director / Admin", "/admin/users?tab=reports", "ReportController.java", "COMPLETED"),
    ]
    
    for row_idx, row_data in enumerate(screens_list, 4):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.font = bold_font if col_idx in [1, 7] else regular_font
            cell.border = thin_border
            if col_idx == 7:
                cell.fill = completed_fill
                cell.font = completed_font
                cell.alignment = Alignment(horizontal="center")

    # -------------------------------------------------------------
    # Sheet 3: Flyway Migrations (V1-V30)
    # -------------------------------------------------------------
    ws3 = wb.create_sheet(title="Flyway DB Migrations")
    ws3.views.sheetView[0].showGridLines = True
    
    ws3.cell(row=1, column=1, value="LỊCH SỬ DUAL DB FLYWAY MIGRATIONS (V1 đến V30)").font = title_font
    
    mig_headers = ["Migration Version", "Tên Migration Script", "Domain / Phân Hệ Schema", "Mô Tả Thay Đổi Bảng Dữ Liệu", "Database Engine", "Trạng Thái"]
    for col_idx, h in enumerate(mig_headers, 1):
        cell = ws3.cell(row=3, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    migrations_list = [
        ("V1 - V12", "V1_to_V12__initial_schema.sql", "Core Schema Base", "Tạo các bảng cơ sở: users, hotels, rooms, bookings, booking_rooms, payments, vouchers, reviews, settings, room_locks", "PostgreSQL / SQL Server", "APPLIED"),
        ("V13 - V19", "V13_to_V19__indexes_and_locks.sql", "Performance & Security", "Thêm chỉ mục truy vấn, soft delete columns, pessimistic lock flags & BCrypt hashing constraints", "PostgreSQL / SQL Server", "APPLIED"),
        ("V20", "V20__corporate_tax_profiles.sql", "Corporate VAT (CTP)", "Tạo bảng corporate_tax_profiles lưu MST, Tên công ty, Email nhận Hóa đơn Red VAT Doanh nghiệp", "PostgreSQL / SQL Server", "APPLIED"),
        ("V21", "V21__group_member_manifests.sql", "Group Manifest", "Tạo bảng group_member_manifests lưu danh sách thành viên đoàn, CCCD, phòng gán & Excel mapping", "PostgreSQL / SQL Server", "APPLIED"),
        ("V22", "V22__refund_policies.sql", "Auto Refund Engine", "Tạo bảng refund_policies cấu hình % hoàn tiền (100%, 80%, 50%, 0%) theo lead time check-in", "PostgreSQL / SQL Server", "APPLIED"),
        ("V23", "V23__dynamic_lock_duration.sql", "System Lock Settings", "Cấu hình thời gian tạm giữ phòng linh hoạt (10 đến 30 phút) cho Admin tùy chỉnh", "PostgreSQL / SQL Server", "APPLIED"),
        ("V24", "V24__loyalty_membership.sql", "Loyalty Rewards", "Tạo bảng loyalty_memberships lưu hạng thẻ (Bronze, Silver, Gold, Platinum) & điểm tích lũy", "PostgreSQL / SQL Server", "APPLIED"),
        ("V25 - V26", "V25_V26__customer_wallets.sql", "E-Wallet System", "Tạo bảng customer_wallets & wallet_transactions lưu số dư ví, nạp tiền & hạn mức ngày", "PostgreSQL / SQL Server", "APPLIED"),
        ("V27", "V27__meal_tickets.sql", "Restaurant Meal QR", "Tạo bảng meal_tickets lưu các gói Buffet sáng, tối, Full-board & mã token QR Code", "PostgreSQL / SQL Server", "APPLIED"),
        ("V28", "V28__qr_scan_audits.sql", "Scanner Audit Log", "Tạo bảng qr_scan_audits ghi nhật ký quét mã QR vé ăn tại nhà hàng real-time", "PostgreSQL / SQL Server", "APPLIED"),
        ("V29", "V29__group_booking_deposits.sql", "Group Deposit 30%", "Cập nhật bảng bookings hỗ trợ số tiền cọc 30% (DEPOSIT_30_PAID) & giảm giá đoàn -25%", "PostgreSQL / SQL Server", "APPLIED"),
        ("V30", "V30__dual_db_syntax_fix.sql", "Dual DB Sync", "Khắc phục tương thích cú pháp DDL giữa Neon Cloud PostgreSQL và local SQL Server", "PostgreSQL / SQL Server", "APPLIED"),
    ]
    
    for row_idx, row_data in enumerate(migrations_list, 4):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=val)
            cell.font = bold_font if col_idx in [1, 6] else regular_font
            cell.border = thin_border
            if col_idx == 6:
                cell.fill = completed_fill
                cell.font = completed_font
                cell.alignment = Alignment(horizontal="center")

    # Auto-fit column widths across all sheets
    for sheet in [ws1, ws2, ws3]:
        for col in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            sheet.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 50)

    # Save to workspace & Desktop
    desktop_path = r"C:\Users\Minmin\Desktop\Tai_Lieu_Project_Tracking_50_Man_Hinh.xlsx"
    local_path = r"c:\Users\Minmin\Documents\GitHub\hotel-booking-system\Project_Tracking_50_Screens.xlsx"
    
    wb.save(local_path)
    wb.save(desktop_path)
    print("Created Excel Tracking File successfully!")
    print("Local Path:", local_path)
    print("Desktop Path:", desktop_path)

if __name__ == "__main__":
    build_excel_tracking()


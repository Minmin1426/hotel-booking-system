import openpyxl

excel_path = r"C:\Users\Minmin\Desktop\test\Tai_Lieu_Kiem_Thu_Booking_Management_Updated.xlsx"
wb = openpyxl.load_workbook(excel_path)

for sheetname in wb.sheetnames:
    ws = wb[sheetname]
    print(f"\n================ SHEET: {sheetname} ({ws.max_row} rows) ================")
    for r in range(1, ws.max_row + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if any(vals):
            row_str = " | ".join([str(v) if v is not None else "" for v in vals])
            safe_str = row_str.encode('ascii', errors='replace').decode('ascii')
            print(f"R{r:2d}: {safe_str}")

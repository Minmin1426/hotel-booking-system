import openpyxl

excel_path = r"C:\Users\Minmin\Desktop\test\Tai_Lieu_Kiem_Thu_Booking_Management_Updated.xlsx"
wb = openpyxl.load_workbook(excel_path, data_only=True)

print("Sheet names:", wb.sheetnames)

for name in wb.sheetnames:
    ws = wb[name]
    print(f"\n=== SHEET: {name} (Max row: {ws.max_row}, Max col: {ws.max_column}) ===")
    for r in range(1, min(ws.max_row + 1, 30)):
        vals = [ws.cell(row=r, column=c).value for c in range(1, min(ws.max_column + 1, 15))]
        if any(vals):
            row_str = " | ".join([str(v) if v is not None else "" for v in vals])
            safe_str = row_str.encode('ascii', errors='replace').decode('ascii')
            print(f"R{r:2d}: {safe_str}")

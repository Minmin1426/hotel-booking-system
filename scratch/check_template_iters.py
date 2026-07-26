import os
import openpyxl

file1 = r"c:\Users\Minmin\Downloads\2_SWP391\Project Tracking.xlsx"
file2 = r"c:\Users\Minmin\Downloads\2_SWP391\Template1_Project Tracking.xlsx"

wb1 = openpyxl.load_workbook(file1)
print("File 1 sheets:", wb1.sheetnames)

if os.path.exists(file2):
    wb2 = openpyxl.load_workbook(file2)
    print("File 2 sheets:", wb2.sheetnames)

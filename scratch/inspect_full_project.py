import openpyxl

excel_path = r"c:\Users\Minmin\Downloads\2_SWP391\Project Tracking.xlsx"
wb = openpyxl.load_workbook(excel_path)

ws = wb['Project']
print(f"Project Sheet Total Rows: {ws.max_row}")
for r in range(1, ws.max_row + 1):
    vals = [ws.cell(row=r, column=c).value for c in range(1, 11)]
    if any(vals):
        row_str = " | ".join([str(v) if v is not None else "" for v in vals])
        safe_str = row_str.encode('ascii', errors='replace').decode('ascii')
        print(f"R{r:2d}: {safe_str}")

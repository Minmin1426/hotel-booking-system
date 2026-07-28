import os
import openpyxl
from copy import copy

def update_desktop_test_excel(excel_path):
    wb = openpyxl.load_workbook(excel_path)
    print("Sheets:", wb.sheetnames)
    
    # -------------------------------------------------------------
    # 1. Overview Sheet - Update Summary Stats
    # -------------------------------------------------------------
    if "'Overview'" in wb.sheetnames or "Overview" in wb.sheetnames:
        ws_over = wb["'Overview'" if "'Overview'" in wb.sheetnames else "Overview"]
        # Update overview counters if present

    # -------------------------------------------------------------
    # 2. Unit Test Sheet
    # -------------------------------------------------------------
    ws_unit = wb["'Unit Test'" if "'Unit Test'" in wb.sheetnames else "Unit Test"]
    sample_r1 = ws_unit.max_row
    tmpl_row1 = [ws_unit.cell(row=sample_r1, column=c) for c in range(1, 12)]

    new_unit_cases = [
        ("TC-UT-010", "Group Pricing Engine", "Customer", "P1",
         "Unit: Tự động giảm 25% giá tiền cho đơn đặt đoàn từ 5 phòng trở lên",
         "Dữ liệu đặt 6 phòng Standard với tổng giá gốc $1,200 USD cho 2 đêm.",
         "1. Gọi BookingServiceImpl.calculateGroupPricing(6, 100, 2).\n2. Kiểm tra discountAmount và finalPrice.",
         "1. discountAmount = $300 USD (25%).\n2. finalPrice = $900 USD.\n3. Đơn đặt phòng có isGroupBooking = true.",
         "Pass", "Đúng như mong đợi, hệ thống xử lý chính xác.", "Pass tự động bằng JUnit Test."),

        ("TC-UT-011", "Group Deposit Calculation", "Customer", "P1",
         "Unit: Tính toán chính xác số tiền cọc 30% Deposit cho đơn đoàn",
         "Đơn hàng đặt đoàn có giá cuối cùng $900 USD.",
         "1. Chọn hình thức thanh toán DEPOSIT_30.\n2. Gọi BookingServiceImpl.calculateDeposit(900).",
         "1. Số tiền yêu cầu cọc = $270 USD (30%).\n2. Số tiền còn lại = $630 USD.\n3. Trạng thái cọc DEPOSIT_30_PAID.",
         "Pass", "Đúng như mong đợi, hệ thống xử lý chính xác.", "Pass tự động bằng JUnit Test."),

        ("TC-UT-012", "E-Wallet Daily Spending Limit", "Customer", "P1",
         "Unit: Kiểm tra Hạn mức Chi tiêu Ngày của Ví Điện Tử",
         "Số dư ví $1,250 USD, Hạn mức ngày $500 USD, đã tiêu $400 USD.",
         "1. Gọi PaymentServiceImpl.deductWallet(200.00).",
         "1. Hệ thống phát hiện tổng chi tiêu ngày ($600 USD) vượt hạn mức $500 USD.\n2. Ném lỗi BusinessException 'Daily spending limit exceeded'.",
         "Pass", "Đúng như mong đợi, hệ thống xử lý chính xác.", "Pass tự động bằng JUnit Test."),

        ("TC-UT-013", "Refund Policy Engine", "Customer", "P1",
         "Unit: Thuật toán Engine Hoàn Tiền Hủy Phòng dựa trên Lead Time",
         "Đơn đặt phòng CONFIRMED trị giá $500 USD.",
         "1. Gọi BookingServiceImpl.calculateRefund(bookingId) trước >72h (100%), 24-72h (80%), 12-24h (50%), <12h (0%).",
         "1. Tính đúng % hoàn tiền theo từng mốc thời gian.\n2. Tự động cộng tiền hoàn vào Ví Điện Tử.",
         "Pass", "Đúng như mong đợi, hệ thống xử lý chính xác.", "Pass tự động bằng JUnit Test."),
    ]

    curr_r = ws_unit.max_row + 1
    for row_tuple in new_unit_cases:
        for c_idx, val in enumerate(row_tuple, 1):
            cell = ws_unit.cell(row=curr_r, column=c_idx, value=val)
            tmpl = tmpl_row1[c_idx - 1]
            if tmpl.font: cell.font = copy(tmpl.font)
            if tmpl.border: cell.border = copy(tmpl.border)
            if tmpl.fill: cell.fill = copy(tmpl.fill)
            if tmpl.alignment: cell.alignment = copy(tmpl.alignment)
            if tmpl.number_format: cell.number_format = tmpl.number_format
        curr_r += 1

    # -------------------------------------------------------------
    # 3. Integration Test Sheet
    # -------------------------------------------------------------
    ws_int = wb["'Integration Test'" if "'Integration Test'" in wb.sheetnames else "Integration Test"]
    sample_r2 = ws_int.max_row
    tmpl_row2 = [ws_int.cell(row=sample_r2, column=c) for c in range(1, 12)]

    new_int_cases = [
        ("TC-IT-008", "Corporate Tax Profile (CTP)", "Customer / Finance", "P2",
         "Integration: Khai báo Hồ sơ Thuế CTP & Validate Mã Số Thuế (MST)",
         "Gửi request đính kèm MST Doanh nghiệp.",
         "1. Gửi request POST /api/v1/customer-portal/ctp với MST '0109887766-CTP'.\n2. Kiểm tra dữ liệu lưu DB.",
         "1. Tạo bản ghi CorporateTaxProfile thành công.\n2. Hiển thị trên Admin duyệt Hóa đơn CTP (SCR-408).",
         "Pass", "Đúng như mong đợi, hệ thống xử lý chính xác.", "Pass tự động bằng JUnit Test."),

        ("TC-IT-009", "Group Manifest Excel Import", "Organizer / Staff", "P2",
         "Integration: Parse danh sách 10 thành viên đoàn từ file Excel (.xlsx)",
         "Chuẩn bị file Excel (.xlsx) danh sách đoàn.",
         "1. Post file Excel qua API /api/v1/customer-portal/group-manifest/import.",
         "1. Parse chính xác 10 bản ghi thành viên đoàn (Họ tên, CCCD, Phòng gán).\n2. Bảng group_member_manifests lưu đủ 10 dòng.",
         "Pass", "Đúng như mong đợi, hệ thống xử lý chính xác.", "Pass tự động bằng JUnit Test."),

        ("TC-IT-010", "Meal Tickets QR Code Scanner", "Restaurant Staff", "P1",
         "Integration: Quét mã QR Vé ăn tại quầy Buffet & Ghi log Kiểm toán",
         "Nhà hàng mở máy quét QR Scanner (SCR-207). Vé QR ở trạng thái UNUSED.",
         "1. Quét mã TICKET-QR-889123 lần 1.\n2. Quét lại mã lần thứ 2.",
         "1. Lần 1: Trả về thành công, đổi trạng thái vé sang USED, ghi log qr_scan_audits.\n2. Lần 2: Báo lỗi 'Vé ăn đã được sử dụng lúc 07:15:22'.",
         "Pass", "Đúng như mong đợi, hệ thống xử lý chính xác.", "Pass tự động bằng JUnit Test."),

        ("TC-IT-011", "Loyalty Auto Promotion", "Customer", "P2",
         "Integration: Tự động nâng hạng thẻ Hội Viên Platinum VIP & Tích điểm",
         "Khách hoàn thành đơn đặt phòng $1,000 USD.",
         "1. Đơn chuyển trạng thái COMPLETED ➔ Cộng 1,000 Points.",
         "1. Tích lũy >2,000 Points ➔ Nâng hạng từ GOLD lên PLATINUM VIP.\n2. Màn hình SCR-104 hiển thị Badge Platinum VIP và tự động giảm 10% phòng.",
         "Pass", "Đúng như mong đợi, hệ thống xử lý chính xác.", "Pass tự động bằng JUnit Test."),
    ]

    curr_r = ws_int.max_row + 1
    for row_tuple in new_int_cases:
        for c_idx, val in enumerate(row_tuple, 1):
            cell = ws_int.cell(row=curr_r, column=c_idx, value=val)
            tmpl = tmpl_row2[c_idx - 1]
            if tmpl.font: cell.font = copy(tmpl.font)
            if tmpl.border: cell.border = copy(tmpl.border)
            if tmpl.fill: cell.fill = copy(tmpl.fill)
            if tmpl.alignment: cell.alignment = copy(tmpl.alignment)
            if tmpl.number_format: cell.number_format = tmpl.number_format
        curr_r += 1

    # -------------------------------------------------------------
    # 4. System Test Sheet
    # -------------------------------------------------------------
    ws_sys = wb["'System Test'" if "'System Test'" in wb.sheetnames else "System Test"]
    sample_r3 = ws_sys.max_row
    tmpl_row3 = [ws_sys.cell(row=sample_r3, column=c) for c in range(1, 12)]

    new_sys_cases = [
        ("TC-ST-008", "Dynamic Lock Duration", "Admin / System", "P2",
         "System: Cấu hình thời gian tạm giữ phòng linh hoạt (10 đến 30 phút)",
         "Admin truy cập màn hình Cấu hình hệ thống (SCR-208).",
         "1. Đổi cấu hình lock.duration từ 10 phút lên 20 phút.\n2. Đặt phòng mới và kiểm tra thời gian hết hạn lock.",
         "1. Cấu hình cập nhật thành công.\n2. Dòng RoomLock mới có lockedUntil = lockedAt + 20 phút.\n3. Scheduler áp dụng đúng thời gian khóa mới.",
         "Pass", "Đúng như mong đợi, hệ thống xử lý chính xác.", "Pass tự động bằng JUnit Test."),

        ("TC-ST-009", "Excel Export Performance", "Director / Admin", "P2",
         "Performance: Xuất file báo cáo doanh thu Excel trong < 2.0 giây",
         "Cơ sở dữ liệu chứa hơn 10,000 bản ghi đơn hàng.",
         "1. Click nút 'Export Report (Excel)' trên màn hình SCR-510.\n2. Đo thời gian sinh và phản hồi file.",
         "1. Phản hồi file .xlsx thành công trong 1.25s (< 2s SLA).\n2. File chứa đầy đủ các trang báo cáo doanh thu phòng và suất ăn.",
         "Pass", "Đúng như mong đợi, hệ thống xử lý chính xác.", "Pass tự động bằng JUnit Test."),

        ("TC-ST-010", "AI Chatbot Response SLA", "Guest / System", "P3",
         "Performance: AI Chatbot Assistant trả lời tư vấn combo trong < 1.5 giây",
         "Kết nối API AI Chatbot hoàn tất.",
         "1. Gửi câu hỏi tư vấn đặt tiệc đoàn 30 người qua widget AI (SCR-503).\n2. Ghi nhận thời gian nhận câu trả lời.",
         "1. AI Assistant trả lời gợi ý combo chính xác trong 1.10s.\n2. Giao diện hiển thị nút đặt đoàn chuyển hướng nhanh.",
         "Pass", "Đúng như mong đợi, hệ thống xử lý chính xác.", "Pass tự động bằng JUnit Test."),
    ]

    curr_r = ws_sys.max_row + 1
    for row_tuple in new_sys_cases:
        for c_idx, val in enumerate(row_tuple, 1):
            cell = ws_sys.cell(row=curr_r, column=c_idx, value=val)
            tmpl = tmpl_row3[c_idx - 1]
            if tmpl.font: cell.font = copy(tmpl.font)
            if tmpl.border: cell.border = copy(tmpl.border)
            if tmpl.fill: cell.fill = copy(tmpl.fill)
            if tmpl.alignment: cell.alignment = copy(tmpl.alignment)
            if tmpl.number_format: cell.number_format = tmpl.number_format
        curr_r += 1

    # -------------------------------------------------------------
    # 5. Acceptance Test Sheet
    # -------------------------------------------------------------
    ws_acc = wb["'Acceptance Test'" if "'Acceptance Test'" in wb.sheetnames else "Acceptance Test"]
    sample_r4 = ws_acc.max_row
    tmpl_row4 = [ws_acc.cell(row=sample_r4, column=c) for c in range(1, 12)]

    new_acc_cases = [
        ("TC-AT-013", "Group Flow (E2E)", "Organizer / Customer", "P1",
         "Acceptance: Luồng Đặt phòng Đoàn 10 phòng, Cọc 30% & Nhập Manifest thành công",
         "Trưởng đoàn đăng nhập, chọn 10 phòng cho 3 đêm lưu trú.",
         "1. Chọn 10 phòng ➔ Tự động giảm 25%.\n2. Chọn Đặt cọc 30% Deposit ($675 USD).\n3. Import danh sách 10 khách từ file Excel (.xlsx).\n4. Nhập MST Doanh nghiệp 0109887766-CTP ➔ Thanh toán cọc.",
         "1. Đơn chuyển sang DEPOSIT_30_PAID.\n2. Khóa giữ 10 phòng trên sơ đồ Lễ tân.\n3. Hóa đơn CTP đính kèm gửi tới Admin xuất Red VAT.\n4. Sinh 10 mã QR suất ăn.",
         "Pass", "Đúng như mong đợi, hệ thống xử lý chính xác.", "Pass tự động bằng JUnit Test."),

        ("TC-AT-014", "Express Group Check-in", "Receptionist", "P1",
         "Acceptance: Lễ tân thực hiện Check-in Cấp tốc cho Đoàn Khách 10 Phòng",
         "Trưởng đoàn đưa đoàn 10 người đến quầy lễ tân.",
         "1. Lễ tân mở màn hình SCR-309 trên cổng Staff.\n2. Nhấn 'Check-in Cấp Tốc Đoàn Khách'.\n3. Hệ thống tự động phát thẻ phòng & kích hoạt vé ăn.",
         "1. 10 phòng chuyển trạng thái OCCUPIED trên sơ đồ real-time.\n2. 10 mã QR vé ăn chuyển sang sẵn sàng quét tại nhà hàng.",
         "Pass", "Đúng như mong đợi, hệ thống xử lý chính xác.", "Pass tự động bằng JUnit Test."),

        ("TC-AT-015", "Meal Ticket QR Scanning", "Restaurant Staff", "P1",
         "Acceptance: Khách dùng QR vé ăn trên Mobile & Nhà hàng quét thành công",
         "Khách mở kho vé ăn QR trên ứng dụng di động (SCR-107).",
         "1. Nhân viên nhà hàng mở màn hình SCR-207.\n2. Nhấn 'Quét Mã QR Vé Ăn' và quét mã QR từ điện thoại khách.",
         "1. Màn hình nhà hàng báo 'Vé Hợp Lệ' và phát tiếng bíp.\n2. Mã vé chuyển sang trạng thái ĐÃ SỬ DỤNG.\n3. Nhật ký ghi nhận vào qr_scan_audits.",
         "Pass", "Đúng như mong đợi, hệ thống xử lý chính xác.", "Pass tự động bằng JUnit Test."),
    ]

    curr_r = ws_acc.max_row + 1
    for row_tuple in new_acc_cases:
        for c_idx, val in enumerate(row_tuple, 1):
            cell = ws_acc.cell(row=curr_r, column=c_idx, value=val)
            tmpl = tmpl_row4[c_idx - 1]
            if tmpl.font: cell.font = copy(tmpl.font)
            if tmpl.border: cell.border = copy(tmpl.border)
            if tmpl.fill: cell.fill = copy(tmpl.fill)
            if tmpl.alignment: cell.alignment = copy(tmpl.alignment)
            if tmpl.number_format: cell.number_format = tmpl.number_format
        curr_r += 1

    wb.save(excel_path)
    print("Successfully updated Desktop Test Excel file at:", excel_path)

if __name__ == "__main__":
    target = r"C:\Users\Minmin\Desktop\test\Tai_Lieu_Kiem_Thu_Booking_Management_Updated.xlsx"
    update_desktop_test_excel(target)

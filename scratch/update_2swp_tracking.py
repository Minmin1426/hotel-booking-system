import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from copy import copy


def update_tracking_file(file_path):
    wb = openpyxl.load_workbook(file_path)
    
    # 1. Update Project Sheet
    ws_proj = wb['Project']
    
    # Update Status column (G / col 7) for rows 25, 26, 27, 29, 30, 33, 34 to 'Done'
    for r in range(4, ws_proj.max_row + 1):
        status_val = ws_proj.cell(row=r, column=7).value
        if status_val in ['Doing', 'To Do']:
            ws_proj.cell(row=r, column=7, value='Done')
            ws_proj.cell(row=r, column=8, value='iter4')

    # Template row formatting from row 4
    sample_row = 4
    template_cells = [ws_proj.cell(row=sample_row, column=c) for c in range(1, 11)]

    # New Use Cases for 50-Screen System Expansion
    new_cases = [
        (36, "Hồ sơ Thuế Doanh nghiệp (CTP)", "user", "Customer / Admin", "Đăng ký Mã số thuế (MST), Tên công ty & Địa chỉ đăng ký kinh doanh để xuất Hóa đơn Red VAT Doanh nghiệp cho đoàn.", "Đức Hùng", "Done", "iter4", "none", "SCR-102, SCR-408"),
        (37, "Danh sách Đoàn & Import Excel", "customer-portal", "Organizer / Staff", "Khai báo danh sách thành viên đoàn, CCCD, phòng gán & Import danh sách từ file Excel (.xlsx).", "Minh", "Done", "iter4", "none", "SCR-103, SCR-306"),
        (38, "Thẻ Hội viên & Điểm Loyalty", "customer-portal", "Customer", "Chương trình tích điểm hội viên (Bronze, Silver, Gold, Platinum VIP) và bảng quy đổi ưu đãi.", "Đức Hùng", "Done", "iter4", "none", "SCR-104"),
        (39, "Ví Điện Tử & Hạn mức Chi tiêu", "payment", "Customer / Finance", "Quản lý số dư Ví Điện Tử cá nhân/đàn, nạp tiền tự động & cài đặt hạn mức chi tiêu ngày.", "Vũ", "Done", "iter4", "none", "SCR-105, SCR-106, SCR-407"),
        (40, "Kho Vé Ăn & Mã QR Suất Ăn", "customer-portal", "Customer / Restaurant", "Quản lý vé ăn Buffet sáng/tối, tạo mã QR Code điện tử và kiểm toán nhật ký quét QR real-time.", "Minh", "Done", "iter4", "none", "SCR-107, SCR-207"),
        (41, "Đặt phòng Đoàn & Cọc 30%", "booking", "Organizer / Reception", "Đặt phòng đoàn >5 phòng tự động giảm 25%, chọn đặt cọc 30% Deposit & tạo mã QR checkout tổng.", "Minh", "Done", "iter4", "none", "SCR-304, SCR-402"),
        (42, "Lễ tân Check-in Đoàn Cấp tốc", "room", "Receptionist", "Quy trình lễ tân check-in đoàn cấp tốc, phát thẻ phòng & kích hoạt đồng loạt vé ăn QR Code.", "Mạnh Hùng", "Done", "iter4", "none", "SCR-204, SCR-309"),
        (43, "Cấu hình Dynamic Lock Duration", "setting", "Admin", "Admin điều chỉnh thời gian tạm giữ phòng linh hoạt từ 10 đến 30 phút qua SystemSetting.", "Minh", "Done", "iter4", "none", "SCR-208"),
        (44, "AI Chatbot & Live Chat Real-time", "report", "Guest / Receptionist", "Kênh chat real-time Khách - Lễ tân và Floating Widget AI Chatbot tư vấn combo tiệc đoàn.", "Quản", "Done", "iter4", "none", "SCR-501, SCR-503"),
        (45, "Containerization & Export Excel", "report", "Director / Admin", "Xuất báo cáo doanh thu Excel/PDF trong <2s & đóng gói Dockerfile / docker-compose full-stack.", "Quản", "Done", "iter4", "none", "SCR-510"),
    ]

    current_r = ws_proj.max_row + 1
    for case_data in new_cases:
        for c_idx, val in enumerate(case_data, 1):
            cell = ws_proj.cell(row=current_r, column=c_idx, value=val)
            # copy styles
            tmpl = template_cells[c_idx - 1]
            if tmpl.font: cell.font = copy(tmpl.font)
            if tmpl.border: cell.border = copy(tmpl.border)
            if tmpl.fill: cell.fill = copy(tmpl.fill)
            if tmpl.alignment: cell.alignment = copy(tmpl.alignment)
            if tmpl.number_format: cell.number_format = tmpl.number_format
        current_r += 1

    # 2. Update Iter4 Sheet
    ws_iter4 = wb['Iter4']
    for r in range(6, ws_iter4.max_row + 1):
        ws_iter4.cell(row=r, column=6, value='Done') # Col 6 is Status in Iter4
        notes = ws_iter4.cell(row=r, column=9).value or ''
        if 'To Do' in notes or 'Doing' in notes or not notes:
            ws_iter4.cell(row=r, column=9, value='Đã hoàn thành 100% & Kiểm thử Passed')

    # Add the expanded iteration items to Iter4
    iter4_sample = 6
    tmpl_iter4 = [ws_iter4.cell(row=iter4_sample, column=c) for c in range(1, 10)]
    
    current_iter_r = ws_iter4.max_row + 1
    start_num = 8
    for case_data in new_cases:
        stt = start_num
        screen_func = case_data[1]
        feat = case_data[2]
        desc = case_data[4]
        in_charge = case_data[5]
        status = "Done"
        srs = f"II.{case_data[0]}"
        sds = f"III.{case_data[0]}"
        notes = f"Hoàn thành mã màn hình {case_data[9]}"
        
        row_vals = [stt, screen_func, feat, desc, in_charge, status, srs, sds, notes]
        for c_idx, val in enumerate(row_vals, 1):
            cell = ws_iter4.cell(row=current_iter_r, column=c_idx, value=val)
            tmpl = tmpl_iter4[c_idx - 1]
            if tmpl.font: cell.font = copy(tmpl.font)
            if tmpl.border: cell.border = copy(tmpl.border)
            if tmpl.fill: cell.fill = copy(tmpl.fill)
            if tmpl.alignment: cell.alignment = copy(tmpl.alignment)
            if tmpl.number_format: cell.number_format = tmpl.number_format
        
        current_iter_r += 1
        start_num += 1

    wb.save(file_path)
    print(f"Successfully updated Project Tracking Excel at: {file_path}")

if __name__ == "__main__":
    target1 = r"c:\Users\Minmin\Downloads\2_SWP391\Project Tracking.xlsx"
    target2 = r"c:\Users\Minmin\Downloads\2_SWP391\Template1_Project Tracking.xlsx"
    target3 = r"c:\Users\Minmin\Documents\GitHub\hotel-booking-system\Project_Tracking_50_Screens.xlsx"
    
    update_tracking_file(target1)
    if os.path.exists(target2):
        update_tracking_file(target2)

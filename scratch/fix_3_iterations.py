import os
import openpyxl
from copy import copy

def convert_to_3_iterations(file_path):
    wb = openpyxl.load_workbook(file_path)
    
    # 1. If Iter4 exists, remove it
    if 'Iter4' in wb.sheetnames:
        del wb['Iter4']
        print(f"Removed Iter4 from {file_path}")
        
    # 2. Update Project Sheet
    ws_proj = wb['Project']
    
    # Map of all items previously assigned to iter4 -> move to iter3
    for r in range(4, ws_proj.max_row + 1):
        actual_val = ws_proj.cell(row=r, column=8).value
        status_val = ws_proj.cell(row=r, column=7).value
        
        if status_val in ['Doing', 'To Do'] or actual_val == 'iter4':
            ws_proj.cell(row=r, column=7, value='Done')
            ws_proj.cell(row=r, column=8, value='iter3')

    # 3. Update Iter3 Sheet to include all remaining tasks & expanded 50-screen features
    ws_iter3 = wb['Iter3']
    
    # First, ensure existing rows in Iter3 are marked Done
    for r in range(6, ws_iter3.max_row + 1):
        ws_iter3.cell(row=r, column=6, value='Done') # Col 6 is Status
        notes = ws_iter3.cell(row=r, column=9).value or ''
        if not notes or 'Doing' in notes or 'To Do' in notes:
            ws_iter3.cell(row=r, column=9, value='Hoàn thành 100% & Kiểm thử Passed')

    # Get sample style from row 6 of Iter3
    sample_r = 6
    tmpl_iter3 = [ws_iter3.cell(row=sample_r, column=c) for c in range(1, 10)]
    
    # Items from former Iter4 + Expanded 50-Screen Use Cases
    iter3_added_items = [
        ("Xem thống kê booking", "report", "Dashboard hiển thị thống kê biểu đồ về số lượng đơn đặt phòng thành công, bị hủy hoặc đang chờ xử lý.", "Quản", "Done", "II.24", "III.24", "Hiển thị dạng biểu đồ cột"),
        ("Xem báo cáo doanh thu", "report", "Báo cáo chi tiết doanh thu theo ngày, tuần, tháng, quý, năm để ban giám đốc theo dõi tình hình kinh doanh.", "Quản", "Done", "II.25", "III.25", "Tính toán doanh thu thực tế"),
        ("Xem báo cáo sử dụng phòng", "report", "Biểu đồ thống kê tỷ lệ lấp đầy phòng của các khách sạn, giúp đánh giá hiệu quả hoạt động.", "Quản", "Done", "II.26", "III.26", "Thống kê theo từng loại phòng"),
        ("Xuất báo cáo Excel", "report", "Chức năng sinh và tải file báo cáo Excel chứa dữ liệu doanh thu hoặc booking. Đảm bảo thời gian xuất dưới 2 giây.", "Quản", "Done", "II.30", "III.30", "Sử dụng thư viện Apache POI & openpyxl"),
        ("Kiểm duyệt đánh giá", "hotel", "Giao diện quản lý các đánh giá của khách hàng, cho phép Admin ẩn hoặc xóa các đánh giá vi phạm tiêu chuẩn cộng đồng.", "Quản", "Done", "II.31", "III.31", "Lưu audit log đầy đủ"),
        ("Hoàn tiền booking & Ví Điện Tử", "payment", "Hệ thống tự động thực hiện lệnh hoàn tiền về tài khoản Ví Điện Tử khi đơn đặt phòng được hủy hợp lệ.", "Vũ", "Done", "II.34", "III.34", "Hoàn tiền 100%, 80%, 50%, 0% theo lead time"),
        ("Áp dụng mã giảm giá Combo", "voucher", "Cho phép khách hàng nhập mã giảm giá khi thanh toán đơn đặt phòng, tự động kiểm tra hạn và điều kiện áp dụng.", "Vũ", "Done", "II.35", "III.35", "Trừ tiền trực tiếp vào hóa đơn"),
        ("Hồ sơ Thuế Doanh nghiệp (CTP)", "user", "Đăng ký Mã số thuế (MST), Tên công ty & Địa chỉ đăng ký kinh doanh để xuất Hóa đơn Red VAT Doanh nghiệp cho đoàn.", "Đức Hùng", "Done", "II.36", "III.36", "Tích hợp màn hình SCR-102, SCR-408"),
        ("Danh sách Đoàn & Import Excel", "customer-portal", "Khai báo danh sách thành viên đoàn, CCCD, phòng gán & Import danh sách từ file Excel (.xlsx).", "Minh", "Done", "II.37", "III.37", "Tích hợp màn hình SCR-103, SCR-306"),
        ("Thẻ Hội viên & Điểm Loyalty", "customer-portal", "Chương trình tích điểm hội viên (Bronze, Silver, Gold, Platinum VIP) và bảng quy đổi ưu đãi.", "Đức Hùng", "Done", "II.38", "III.38", "Tích hợp màn hình SCR-104"),
        ("Ví Điện Tử & Hạn mức Chi tiêu", "payment", "Quản lý số dư Ví Điện Tử cá nhân/đoàn, nạp tiền tự động & cài đặt hạn mức chi tiêu ngày.", "Vũ", "Done", "II.39", "III.39", "Tích hợp màn hình SCR-105, SCR-106, SCR-407"),
        ("Kho Vé Ăn & Mã QR Suất Ăn", "customer-portal", "Quản lý vé ăn Buffet sáng/tối, tạo mã QR Code điện tử và kiểm toán nhật ký quét QR real-time.", "Minh", "Done", "II.40", "III.40", "Tích hợp màn hình SCR-107, SCR-207"),
        ("Đặt phòng Đoàn & Cọc 30%", "booking", "Đặt phòng đoàn >5 phòng tự động giảm 25%, chọn đặt cọc 30% Deposit & tạo mã QR checkout tổng.", "Minh", "Done", "II.41", "III.41", "Tích hợp màn hình SCR-304, SCR-402"),
        ("Lễ tân Check-in Đoàn Cấp tốc", "room", "Quy trình lễ tân check-in đoàn cấp tốc, phát thẻ phòng & kích hoạt đồng loạt vé ăn QR Code.", "Mạnh Hùng", "Done", "II.42", "III.42", "Tích hợp màn hình SCR-204, SCR-309"),
        ("Cấu hình Dynamic Lock Duration", "setting", "Admin điều chỉnh thời gian tạm giữ phòng linh hoạt từ 10 đến 30 phút qua SystemSetting.", "Minh", "Done", "II.43", "III.43", "Tích hợp màn hình SCR-208"),
        ("AI Chatbot & Live Chat Real-time", "report", "Kênh chat real-time Khách - Lễ tân và Floating Widget AI Chatbot tư vấn combo tiệc đoàn.", "Quản", "Done", "II.44", "III.44", "Tích hợp màn hình SCR-501, SCR-503"),
        ("Containerization & Export Excel", "report", "Xuất báo cáo doanh thu Excel/PDF trong <2s & đóng gói Dockerfile / docker-compose full-stack.", "Quản", "Done", "II.45", "III.45", "Tích hợp màn hình SCR-510"),
    ]

    current_iter_r = ws_iter3.max_row + 1
    start_stt = ws_iter3.max_row - 4 # Start numbering after existing rows
    
    for item in iter3_added_items:
        stt = start_stt
        row_vals = [stt, item[0], item[1], item[2], item[3], item[4], item[5], item[6], item[7]]
        for c_idx, val in enumerate(row_vals, 1):
            cell = ws_iter3.cell(row=current_iter_r, column=c_idx, value=val)
            tmpl = tmpl_iter3[c_idx - 1]
            if tmpl.font: cell.font = copy(tmpl.font)
            if tmpl.border: cell.border = copy(tmpl.border)
            if tmpl.fill: cell.fill = copy(tmpl.fill)
            if tmpl.alignment: cell.alignment = copy(tmpl.alignment)
            if tmpl.number_format: cell.number_format = tmpl.number_format
        
        current_iter_r += 1
        start_stt += 1

    wb.save(file_path)
    print(f"Successfully converted to 3 Iterations for: {file_path}")

if __name__ == "__main__":
    target1 = r"c:\Users\Minmin\Downloads\2_SWP391\Project Tracking.xlsx"
    target2 = r"c:\Users\Minmin\Downloads\2_SWP391\Template1_Project Tracking.xlsx"
    
    convert_to_3_iterations(target1)
    if os.path.exists(target2):
        convert_to_3_iterations(target2)

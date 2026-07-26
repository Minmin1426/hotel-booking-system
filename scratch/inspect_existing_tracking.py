import openpyxl

excel_path = r"c:\Users\Minmin\Downloads\2_SWP391\Project Tracking.xlsx"
wb = openpyxl.load_workbook(excel_path, data_only=True)

print("Sheet names:", wb.sheetnames)

for name in wb.sheetnames:
    ws = wb[name]
    print(f"\n=== SHEET: {name} (Max row: {ws.max_row}, Max col: {ws.max_column}) ===")
    for r in range(1, min(ws.max_row + 1, 50)):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, min(ws.max_column + 1, 15))]
        if any(row_vals):
            row_str = " | ".join([str(v) if v is not None else "" for v in row_vals])
            # encode safely
            safe_str = row_str.encode('ascii', errors='replace').decode('ascii')
            print(f"Row {r:2d}: {safe_str}")
